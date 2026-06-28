"""Audit: Pending C-grade Uploading.

Business rule:
    Movement 961 = offloading (C-grade taken off).
    Movement 969 = uploading (C-grade put back).
    For every Reference, the count of 969 must equal the count of 961.
    Any shortfall is a PENDING UPLOAD that must be escalated to the
    responsible team.

Pre-processing steps before the check:
    1. Drop rows where 'Base Unit of Measure' = 'KG' (not in scope).
    2. Yellow rows in the 'Quantity' column are subtotals of the
       non-yellow rows directly above them. If the subtotal is 0 or
       positive, the group is balanced — drop the group and its
       subtotal. Only groups with a negative subtotal (pending
       uploads) remain.
    3. Drop rows whose 'Material Description' contains 'sticker'
       (out of scope for this audit).

Outputs:
    data                       - the pending 961 detail rows
                                 (rows to escalate)
    reports['pending_summary'] - per Reference: count_961, count_969,
                                 shortfall
    reports['all_references']  - per Reference 961 vs 969 counts
                                 (full picture, for cross-check)
    notes                      - one-paragraph audit finding
"""

import pandas as pd
from openpyxl import load_workbook


YELLOW_PREFIX = "FFFFEE"


def _yellow_flags(file_path: str, sheet_name: str,
                  qty_header: str = "Quantity") -> list[bool]:
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    qcol = next((c.column for c in header_row if c.value == qty_header), None)
    if qcol is None:
        return []
    flags = []
    for r in range(2, ws.max_row + 1):
        fg = ws.cell(row=r, column=qcol).fill.fgColor
        rgb = str(fg.rgb) if fg and fg.rgb else ""
        flags.append(rgb.upper().startswith(YELLOW_PREFIX))
    return flags


def run(df: pd.DataFrame, file_path: str = None, sheet_name: str = None, **_) -> dict:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    flags = _yellow_flags(file_path, sheet_name) if file_path and sheet_name else []
    if len(flags) < len(df):
        flags = flags + [False] * (len(df) - len(flags))
    df["_yellow"] = flags[: len(df)]

    n0 = len(df)

    # Step 1 — drop KG rows
    uom = df["Base Unit of Measure"].astype(str).str.strip().str.upper()
    df = df.loc[uom != "KG"].reset_index(drop=True)
    n1 = len(df)

    # Step 2 — drop balanced yellow-subtotal groups (subtotal >= 0)
    keep = [True] * len(df)
    group_idx: list[int] = []
    for i, row in df.iterrows():
        group_idx.append(i)
        if row["_yellow"]:
            try:
                qty = float(row["Quantity"])
            except (TypeError, ValueError):
                qty = 0.0
            if qty >= 0:
                for j in group_idx:
                    keep[j] = False
            group_idx = []
    df = df.loc[keep].reset_index(drop=True)
    n2 = len(df)

    # Step 3 — drop 'sticker' descriptions
    desc = df["Material Description"].astype(str).str.lower()
    df = df.loc[~desc.str.contains("sticker", na=False)].reset_index(drop=True)
    n3 = len(df)

    # Drop the empty yellow-subtotal artifact rows from the final detail —
    # they were only needed to anchor the grouping in step 2.
    detail = df.loc[~df["_yellow"]].drop(columns=["_yellow"]).reset_index(drop=True)

    # Per-Reference 961 vs 969 counts (computed from detail rows only)
    mv = pd.to_numeric(detail["Movement Type"], errors="coerce")
    work = detail.assign(_mt=mv)
    per_ref = (
        work.groupby("Reference", dropna=False)["_mt"]
        .agg(
            count_961=lambda s: int((s == 961).sum()),
            count_969=lambda s: int((s == 969).sum()),
        )
        .reset_index()
    )
    per_ref["shortfall"] = (per_ref["count_961"] - per_ref["count_969"]).clip(lower=0)
    pending_summary = (
        per_ref.loc[per_ref["shortfall"] > 0]
        .sort_values("shortfall", ascending=False)
        .reset_index(drop=True)
    )

    # Pending detail = 961 rows whose Reference is in the pending list
    pending_refs = set(pending_summary["Reference"].tolist())
    pending_detail = detail.loc[
        (mv == 961) & (detail["Reference"].isin(pending_refs))
    ].reset_index(drop=True)

    output_cols = [
        "Material", "Material Description", "Material Document",
        "Storage Location", "Posting Date", "Batch", "Quantity",
        "Base Unit of Measure", "Movement Type", "User Name",
        "Reference", "Order", "Customer",
    ]
    pending_detail = pending_detail[[c for c in output_cols if c in pending_detail.columns]]

    total_shortfall = int(pending_summary["shortfall"].sum()) if len(pending_summary) else 0
    notes = (
        f"Audit: Pending C-grade Uploading. "
        f"Source rows: {n0}; after KG drop: {n1}; after balanced-group drop: {n2}; "
        f"after sticker drop: {n3}. "
        f"{len(pending_summary)} Reference(s) have pending uploads — "
        f"total of {total_shortfall} 961 movement(s) await a matching 969 upload."
    )

    return {
        "data": pending_detail,
        "reports": {},
        "notes": notes,
    }
